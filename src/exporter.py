"""
Excel 导出模块
==============
功能：将匹配结果写入 Excel 文件，生成可点击的带价签超链接。

⚠️ 不使用 =HYPERLINK() 公式，而是通过 openpyxl 原生超链接机制写入：
      - cell.value = 显示文本（"点击查看 ¥0.85"）
      - cell.hyperlink = 商品 URL
   这种方式避免了 URL 中 & 等特殊字符导致 XML 解析失败的问题。
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console

from src.taobao_client import TaobaoItem

logger = logging.getLogger(__name__)
console = Console()

OUTPUT_COLUMNS = [
    "匹配价格(元)",
    "匹配店铺名",
    "淘宝链接（带价签）",
    "匹配状态",
]

COLUMN_WIDTHS = {
    "匹配价格(元)": 14,
    "匹配店铺名": 24,
    "淘宝链接（带价签）": 40,
    "匹配状态": 16,
}


def _normalize_url(url: str) -> str:
    url = url.strip()
    if url.startswith("//"):
        return "https:" + url
    if not url.startswith("http"):
        return "https://" + url
    return url


def write_results(original_path: str,
                  original_df: pd.DataFrame,
                  matched_results: list[dict]) -> str:
    """
    将匹配结果写入 `[原文件名]_filled.xlsx`。

    超链接通过 openpyxl 原生机制写入（cell.hyperlink），
    不是手写 =HYPERLINK() 公式，更稳定可靠。
    """
    original_path = Path(original_path)
    output_path = original_path.parent / f"{original_path.stem}_filled.xlsx"

    # ---- 收集数据 ----
    prices: list[Optional[float]] = []
    shops: list[str] = []
    statuses: list[str] = []
    link_data: dict[int, tuple[str, str]] = {}  # 行号 -> (显示文本, URL)

    for idx, result in enumerate(matched_results):
        item: Optional[TaobaoItem] = result.get("item")
        status: str = result.get("status", "未找到商品")

        if item and status == "成功":
            price_val = round(item.price, 2)
            prices.append(price_val)
            shops.append(item.shop)
            statuses.append("成功")
            link_data[idx] = (
                f"点击查看 ¥{price_val:.2f}",
                _normalize_url(item.url),
            )
        else:
            prices.append(None)
            shops.append("")
            statuses.append(status)

    # ---- DataFrame 写出（链接列留空，稍后覆盖）----
    out_df = original_df.copy()
    out_df["匹配价格(元)"] = prices
    out_df["匹配店铺名"] = shops
    out_df["淘宝链接（带价签）"] = ""
    out_df["匹配状态"] = statuses

    out_df.to_excel(output_path, index=False, engine="openpyxl")

    # ---- openpyxl 修补：写入超链接 + 格式化 ----
    wb = load_workbook(output_path)
    ws = wb.active

    # 定位输出列
    col_map = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=ci).value
        if v in OUTPUT_COLUMNS:
            col_map[v] = ci

    # ---- 链接列：原生超链接 ----
    link_col = col_map.get("淘宝链接（带价签）")
    if link_col:
        link_font = Font(color="0563C1", underline="single")
        for row_idx in range(2, ws.max_row + 1):
            data_row = row_idx - 2
            cell = ws.cell(row=row_idx, column=link_col)
            info = link_data.get(data_row)
            if info:
                display_text, url = info
                cell.value = display_text
                cell.hyperlink = url          # openpyxl 原生超链接，不用公式
                cell.font = link_font
            else:
                cell.value = None

    # ---- 表头样式 ----
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    for name, ci in col_map.items():
        c = ws.cell(row=1, column=ci)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align

    # ---- 价格列数字格式 ----
    price_col = col_map.get("匹配价格(元)")
    if price_col:
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=price_col)
            if c.value is not None:
                c.number_format = "0.00"

    # ---- 列宽 ----
    for name, ci in col_map.items():
        ws.column_dimensions[get_column_letter(ci)].width = COLUMN_WIDTHS.get(name, 12)

    # ---- 冻结首行 ----
    ws.freeze_panes = "A2"

    wb.save(output_path)
    wb.close()

    console.print(f"[bold green]✅ 结果已保存至: {output_path}[/bold green]")
    return str(output_path)